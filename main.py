"""
Programme principal - Scraper complet
Supporte : Prada, Kleman, La Bottega Gardiane
Usage : python main.py <URL> [Homme|Femme]
"""

import sys
import time
import os
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_FILE = "etudes_de_cas.xlsx"

GENDER_KEYWORDS = {
    "Homme":   ["mens", "homme", "man", "uomo", "masculin"],
    "Femme":   ["womens", "femme", "woman", "donna", "feminin"],
    "Unisexe": ["unisex", "unisexe"],
}
TYPE_KEYWORDS = {
    "Shoes":     ["chaussure", "shoe", "basket", "sneaker", "derby", "mocassin", "boot", "scarpe", "derbies", "santiag", "western"],
    "Bag":       ["sac", "bag", "pochette", "handbag", "borsa"],
    "Clothing":  ["veste", "manteau", "robe", "pantalon", "jacket", "coat", "dress", "pull", "denim"],
    "Accessory": ["ceinture", "belt", "foulard", "scarf", "chapeau", "hat"],
}

def guess_gender(text):
    text_lower = text.lower()
    for genre, keywords in GENDER_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            return genre
    return None

def guess_type(text):
    text_lower = text.lower()
    for product_type, keywords in TYPE_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            return product_type
    return None

def accepter_cookies(page):
    selecteurs = [
        "button:has-text('Accepter')", "button:has-text('Accept')",
        "button:has-text('Tout accepter')", "button:has-text('Accept all')",
        "#onetrust-accept-btn-handler", ".js-accept-cookies",
        "button:has-text('OK')", "button:has-text('Continuer')",
    ]
    for selector in selecteurs:
        try:
            el = page.query_selector(selector)
            if el and el.is_visible():
                el.click()
                print("  Cookies acceptes automatiquement !")
                time.sleep(1)
                return True
        except Exception:
            continue
    return False

# ─────────────────────────────────────────────
# Scraper produit (onglet 1)
# ─────────────────────────────────────────────
def scraper_produit(page, url):
    titre = None
    try:
        titre = page.eval_on_selector("h1", "el => el.innerText.trim()")
    except Exception:
        pass
    if not titre:
        titre = page.title().split("|")[0].strip()

    page_text = page.inner_text("body") if page.query_selector("body") else ""

    gender = None
    try:
        datalayer = page.evaluate("() => JSON.stringify(window.dataLayer)")
        if datalayer:
            gender = guess_gender(datalayer)
    except Exception:
        pass
    if not gender:
        gender = guess_gender(page_text)

    type_produit = guess_type(page_text)
    return {"titre": titre, "gender": gender, "type": type_produit, "url": url}

# ─────────────────────────────────────────────
# Scraper PRADA
# ─────────────────────────────────────────────
def lire_tableau_prada(page):
    data = {}
    rows = page.query_selector_all("tr.size-table__row")
    for row in rows:
        header = row.query_selector("th.size-table__table-header")
        cells = row.query_selector_all("td.size-table__data")
        if header and cells:
            key = header.inner_text().strip()
            values = [c.inner_text().strip().replace(" cm", "") for c in cells]
            data[key] = values
    return data

def scraper_guide_prada(page):
    print("  Ouverture du guide de taille Prada...")
    try:
        page.click("[data-element='size-guide-trigger']", timeout=10000)
        time.sleep(3)
    except Exception:
        print("  Bouton guide introuvable")
        return [], None
    try:
        page.wait_for_selector("table.size-table__table", timeout=8000)
    except Exception:
        return [], None

    print("  Lecture EU...")
    data_eu = lire_tableau_prada(page)
    taille_marque = data_eu.get("Taille Prada", [])
    eu = data_eu.get("Europe", [])
    cm = data_eu.get("Pied", [])

    print("  Lecture UK...")
    try:
        page.select_option("select[name='select country']", value="gb")
        time.sleep(2)
        data_uk = lire_tableau_prada(page)
        uk = data_uk.get("Royaume-Uni", [])
    except Exception:
        uk = []

    print("  Lecture US...")
    try:
        page.select_option("select[name='select country']", value="us")
        time.sleep(2)
        data_us = lire_tableau_prada(page)
        us = data_us.get("États-Unis", [])
    except Exception:
        us = []

    lignes = []
    for i in range(len(taille_marque)):
        lignes.append({
            "taille_marque": taille_marque[i] if i < len(taille_marque) else None,
            "taille_eu":     eu[i]            if i < len(eu)            else None,
            "taille_uk":     uk[i]            if i < len(uk)            else None,
            "taille_us":     us[i]            if i < len(us)            else None,
            "taille_it":     None,
            "longueur_cm":   cm[i]            if i < len(cm)            else None,
        })
    return lignes, "Prada"

# ─────────────────────────────────────────────
# Scraper KLEMAN
# ─────────────────────────────────────────────
def scraper_guide_kleman(page, gender="Homme"):
    print("  Ouverture du guide de taille Kleman...")
    try:
        page.click("text=Sélectionner une taille", timeout=8000)
        time.sleep(2)
    except Exception:
        pass
    try:
        page.click("text=Guide des tailles", timeout=8000)
        time.sleep(2)
    except Exception:
        print("  Lien Guide des tailles introuvable")
        return [], None
    try:
        page.wait_for_selector(".size-guide-table__content__row", timeout=8000)
    except Exception:
        return [], None

    titre_tableau = "Pointures Homme" if gender == "Homme" else "Pointures Femmes"
    tableaux = page.query_selector_all(".size-guide-table")
    tableau_cible = None
    for tableau in tableaux:
        try:
            titre_el = tableau.evaluate_handle("el => el.previousElementSibling")
            titre_text = page.evaluate("el => el ? el.innerText : ''", titre_el)
            if titre_tableau.lower() in titre_text.lower():
                tableau_cible = tableau
                break
        except Exception:
            continue
    if not tableau_cible and tableaux:
        tableau_cible = tableaux[0]
    if not tableau_cible:
        return [], None

    lignes_html = tableau_cible.query_selector_all(".size-guide-table__content__row")
    lignes = []
    for row in lignes_html[1:]:
        items = row.query_selector_all(".size-guide-table__content__item")
        valeurs = []
        for item in items:
            style = item.get_attribute("style") or ""
            x_show = item.get_attribute("x-show") or ""
            if "display: none" in style or "Pouces" in x_show:
                continue
            valeurs.append(item.inner_text().strip())
        if len(valeurs) >= 4:
            lignes.append({
                "taille_marque": valeurs[0],
                "taille_eu":     valeurs[0],
                "taille_uk":     valeurs[1],
                "taille_us":     valeurs[2],
                "taille_it":     None,
                "longueur_cm":   valeurs[3],
            })
    return lignes, "Kleman"

# ─────────────────────────────────────────────
# Scraper LA BOTTEGA GARDIANE
# ─────────────────────────────────────────────
def scraper_guide_gardiane(page, gender="Homme"):
    print("  Ouverture du guide de taille La Bottega Gardiane...")
    try:
        page.click("[data-trigger-size-guide]", timeout=10000)
        time.sleep(3)
    except Exception:
        print("  Bouton guide introuvable")
        return [], None
    try:
        page.wait_for_selector(".size-guide__table-right-col", timeout=8000)
    except Exception:
        return [], None

    titre_cherche = "POINTURES HOMME" if gender == "Homme" else "POINTURES FEMME"
    print(f"  Lecture tableau : {titre_cherche}")

    tableaux = page.query_selector_all("#splide05-slide01 .size-guide__table")
    tableau_cible = None
    for tableau in tableaux:
        try:
            header = tableau.query_selector(".size-guide__table-cell.is--header")
            if header and titre_cherche in header.inner_text().strip().upper():
                tableau_cible = tableau
                break
        except Exception:
            continue

    if not tableau_cible:
        tous = page.query_selector_all(".size-guide__table")
        for tableau in tous:
            try:
                header = tableau.query_selector(".size-guide__table-cell.is--header")
                if header and titre_cherche in header.inner_text().strip().upper():
                    tableau_cible = tableau
                    break
            except Exception:
                continue

    if not tableau_cible:
        tous = page.query_selector_all(".size-guide__table")
        tableau_cible = tous[0] if tous else None

    if not tableau_cible:
        return [], None

    colonnes = tableau_cible.query_selector_all(".size-guide__table-right-col")
    lignes = []
    for col in colonnes:
        cells = col.query_selector_all(".size-guide__table-cell")
        valeurs = [c.inner_text().strip().replace("cm", "").strip() for c in cells]
        if len(valeurs) >= 5:
            cm_val = valeurs[0].replace(",", ".") if valeurs[0] else None
            fr_val = valeurs[1].replace(",", ".") if valeurs[1] else None
            uk_val = valeurs[2].replace(",", ".") if valeurs[2] else None
            us_val = valeurs[3].replace(",", ".") if valeurs[3] else None
            it_val = valeurs[4].replace(",", ".") if valeurs[4] else None
            if fr_val and fr_val != "-":
                lignes.append({
                    "taille_marque": fr_val,
                    "taille_eu":     fr_val,
                    "taille_uk":     uk_val,
                    "taille_us":     us_val,
                    "taille_it":     it_val,
                    "longueur_cm":   cm_val,
                })
    return lignes, "La Bottega Gardiane"

# ─────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────
def initialiser_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Pages produit"
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    headers = ["Nom Produit", "Gender", "Type", "URL", "Guide de taille"]
    widths  = [40, 12, 12, 55, 16]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws1.column_dimensions[get_column_letter(col)].width = w
    wb.create_sheet("Guides de taille")
    wb.save(EXCEL_FILE)
    print(f"  Fichier '{EXCEL_FILE}' cree !")

def ajouter_onglet1(produit, guide_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Pages produit"]
    next_row = ws.max_row + 1
    fill = PatternFill("solid", start_color="D6E4F0" if next_row % 2 == 0 else "FFFFFF")
    normal_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    valeurs = [produit["titre"], produit["gender"], produit["type"], produit["url"], guide_id]
    for col, val in enumerate(valeurs, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.alignment = center if col in [2, 3, 5] else Alignment(vertical="center")
    wb.save(EXCEL_FILE)
    print(f"  Onglet 1 mis a jour (ligne {next_row})")

def get_prochain_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == "Guide de taille" and row[1] and isinstance(row[1], int):
            max_id = max(max_id, row[1])
    return max_id + 1

def ajouter_onglet2(lignes, url, brand, guide_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Guides de taille"]

    next_row = 1
    for row in ws.iter_rows():
        if any(cell.value is not None for cell in row):
            next_row = row[0].row + 1
    if next_row > 1:
        next_row += 1

    center     = Alignment(horizontal="center", vertical="center")
    blue_fill  = PatternFill("solid", start_color="00B0F0")
    dark_fill  = PatternFill("solid", start_color="1F4E79")
    cyan_fill  = PatternFill("solid", start_color="00FFFF")
    white_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    bold_font  = Font(bold=True, name="Arial", size=10)
    norm_font  = Font(name="Arial", size=10)

    ws.cell(row=next_row, column=1, value="Guide de taille").font = white_font
    ws.cell(row=next_row, column=1).fill = dark_fill
    ws.cell(row=next_row, column=1).alignment = center
    ws.cell(row=next_row, column=2, value=guide_id).font = bold_font
    ws.cell(row=next_row, column=2).fill = cyan_fill
    ws.cell(row=next_row, column=2).alignment = center
    ws.cell(row=next_row, column=3, value="URL").font = bold_font
    ws.cell(row=next_row, column=3).alignment = center
    ws.cell(row=next_row, column=4, value=url).font = norm_font

    row_h = next_row + 2
    ws.cell(row=row_h, column=1, value="Systemes metriques").font = white_font
    ws.cell(row=row_h, column=1).fill = dark_fill
    ws.cell(row=row_h, column=1).alignment = center
    ws.cell(row=row_h, column=2, value="").fill = blue_fill
    for i in range(len(lignes)):
        cell = ws.cell(row=row_h, column=i + 3, value=f"Taille {i+1}")
        cell.font = white_font
        cell.fill = dark_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(i + 3)].width = 10

    has_it = any(l.get("taille_it") for l in lignes)
    rows_data = [
        (brand,         brand, "taille_marque"),
        ("Europe",      "EU",  "taille_eu"),
        ("Royaume-Uni", "UK",  "taille_uk"),
        ("Etats-Unis",  "US",  "taille_us"),
    ]
    if has_it:
        rows_data.append(("Italie", "IT", "taille_it"))

    for offset, (label, short, key) in enumerate(rows_data, 1):
        r = row_h + offset
        ws.cell(row=r, column=1, value=label).font = norm_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2, value=short).font = white_font
        ws.cell(row=r, column=2).fill = blue_fill
        ws.cell(row=r, column=2).alignment = center
        for i, ligne in enumerate(lignes):
            cell = ws.cell(row=r, column=i + 3, value=ligne.get(key))
            cell.font = norm_font
            cell.alignment = center

    r_cm = row_h + len(rows_data) + 1
    ws.cell(row=r_cm, column=1, value="Longueur pied").font = norm_font
    ws.cell(row=r_cm, column=1).alignment = center
    for i, ligne in enumerate(lignes):
        val = f"{ligne['longueur_cm']} cm" if ligne.get("longueur_cm") else None
        cell = ws.cell(row=r_cm, column=i + 3, value=val)
        cell.font = norm_font
        cell.alignment = center

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50

    wb.save(EXCEL_FILE)
    print(f"  Onglet 2 mis a jour (ID guide = {guide_id})")

# ─────────────────────────────────────────────
# Programme principal
# ─────────────────────────────────────────────
def main(url, gender="Homme"):
    initialiser_excel()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="fr-FR",
        )
        page = context.new_page()
        page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)

        print(f"\n Chargement de la page...")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        print(" Attente du chargement JavaScript...")
        time.sleep(4)
        accepter_cookies(page)
        time.sleep(2)

        # Scraper produit
        print("\n Scraping du produit...")
        produit = scraper_produit(page, url)
        print(f"  Titre  : {produit['titre']}")
        print(f"  Gender : {produit['gender']}")
        print(f"  Type   : {produit['type']}")

        # Scraper guide selon le site
        print("\n Scraping du guide de taille...")
        if "prada.com" in url:
            lignes, brand = scraper_guide_prada(page)
        elif "kleman" in url:
            lignes, brand = scraper_guide_kleman(page, gender)
        elif "labottegardiane" in url or "bottega" in url:
            lignes, brand = scraper_guide_gardiane(page, gender)
        else:
            print("  Site non supporte (Prada, Kleman, La Bottega Gardiane supportes)")
            lignes, brand = [], None

        browser.close()

    # ID du guide
    wb = load_workbook(EXCEL_FILE)
    ws2 = wb["Guides de taille"]
    guide_id = get_prochain_id(ws2) if lignes else None

    # Export Excel
    print("\n Export vers Excel...")
    ajouter_onglet1(produit, guide_id)
    if lignes:
        ajouter_onglet2(lignes, url, brand, guide_id)

    print(f"\n Termine ! Ouvre '{EXCEL_FILE}' pour voir le resultat.")
    print(f"  Produit  : {produit['titre']}")
    print(f"  Brand    : {brand or 'Non detecte'}")
    print(f"  Guide ID : {guide_id or 'Aucun'}")
    print(f"  Tailles  : {len(lignes)}\n")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("\nUsage : python main.py <URL> [Homme|Femme]")
        print("Exemples :")
        print("  python main.py https://www.prada.com/...")
        print("  python main.py https://kleman-france.com/... Homme")
        print("  python main.py https://www.labottegardiane.com/... Femme\n")
        sys.exit(1)

    url    = sys.argv[1]
    gender = sys.argv[2] if len(sys.argv) > 2 else "Homme"

    if not url.startswith("http"):
        print("\nErreur : l'URL doit commencer par http://\n")
        sys.exit(1)
    try:
        main(url, gender)
    except Exception as e:
        print(f"\nErreur : {e}\n")
        sys.exit(1)
