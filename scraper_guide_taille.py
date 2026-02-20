"""
Scraper de guide de taille
Supporte : Prada, Kleman, La Bottega Gardiane
Usage : python scraper_guide_taille.py <URL> [Homme|Femme]
"""

import sys
import time
import os
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_FILE = "etudes_de_cas.xlsx"

# ─────────────────────────────────────────────
# Cookies
# ─────────────────────────────────────────────
def accepter_cookies(page):
    selecteurs = [
        "button:has-text('Accepter')", "button:has-text('Accept')",
        "button:has-text('Tout accepter')", "button:has-text('Accept all')",
        "#onetrust-accept-btn-handler", ".js-accept-cookies",
        "button:has-text('OK')", "button:has-text('Continuer')",
    ]
    for selector in selecteurs:
        try:
            el = page.query_selector(selector)
            if el and el.is_visible():
                el.click()
                print("  Cookies acceptes automatiquement !")
                time.sleep(1)
                return True
        except Exception:
            continue
    return False

# ─────────────────────────────────────────────
# Scraper PRADA
# ─────────────────────────────────────────────
def lire_tableau_prada(page):
    data = {}
    rows = page.query_selector_all("tr.size-table__row")
    for row in rows:
        header = row.query_selector("th.size-table__table-header")
        cells = row.query_selector_all("td.size-table__data")
        if header and cells:
            key = header.inner_text().strip()
            values = [c.inner_text().strip().replace(" cm", "") for c in cells]
            data[key] = values
    return data

def scraper_guide_prada(page):
    print("  Ouverture du guide de taille Prada...")
    try:
        page.click("[data-element='size-guide-trigger']", timeout=10000)
        time.sleep(3)
    except Exception:
        print("  Bouton guide introuvable")
        return [], None
    try:
        page.wait_for_selector("table.size-table__table", timeout=8000)
    except Exception:
        return [], None

    print("  Lecture EU...")
    data_eu = lire_tableau_prada(page)
    taille_marque = data_eu.get("Taille Prada", [])
    eu = data_eu.get("Europe", [])
    cm = data_eu.get("Pied", [])

    print("  Lecture UK...")
    try:
        page.select_option("select[name='select country']", value="gb")
        time.sleep(2)
        data_uk = lire_tableau_prada(page)
        uk = data_uk.get("Royaume-Uni", [])
    except Exception:
        uk = []

    print("  Lecture US...")
    try:
        page.select_option("select[name='select country']", value="us")
        time.sleep(2)
        data_us = lire_tableau_prada(page)
        us = data_us.get("États-Unis", [])
    except Exception:
        us = []

    lignes = []
    for i in range(len(taille_marque)):
        lignes.append({
            "taille_marque": taille_marque[i] if i < len(taille_marque) else None,
            "taille_eu":     eu[i]            if i < len(eu)            else None,
            "taille_uk":     uk[i]            if i < len(uk)            else None,
            "taille_us":     us[i]            if i < len(us)            else None,
            "taille_it":     None,
            "longueur_cm":   cm[i]            if i < len(cm)            else None,
        })
    return lignes, "Prada"

# ─────────────────────────────────────────────
# Scraper KLEMAN
# ─────────────────────────────────────────────
def scraper_guide_kleman(page, gender="Homme"):
    print("  Ouverture du guide de taille Kleman...")
    try:
        page.click("text=Sélectionner une taille", timeout=8000)
        time.sleep(2)
    except Exception:
        pass
    try:
        page.click("text=Guide des tailles", timeout=8000)
        time.sleep(2)
    except Exception:
        print("  Lien Guide des tailles introuvable")
        return [], None
    try:
        page.wait_for_selector(".size-guide-table__content__row", timeout=8000)
    except Exception:
        return [], None

    titre_tableau = "Pointures Homme" if gender == "Homme" else "Pointures Femmes"
    print(f"  Lecture tableau : {titre_tableau}")

    tableaux = page.query_selector_all(".size-guide-table")
    tableau_cible = None
    for tableau in tableaux:
        try:
            titre_el = tableau.evaluate_handle("el => el.previousElementSibling")
            titre_text = page.evaluate("el => el ? el.innerText : ''", titre_el)
            if titre_tableau.lower() in titre_text.lower():
                tableau_cible = tableau
                break
        except Exception:
            continue
    if not tableau_cible and tableaux:
        tableau_cible = tableaux[0]
    if not tableau_cible:
        return [], None

    lignes_html = tableau_cible.query_selector_all(".size-guide-table__content__row")
    lignes = []
    for row in lignes_html[1:]:
        items = row.query_selector_all(".size-guide-table__content__item")
        valeurs = []
        for item in items:
            style = item.get_attribute("style") or ""
            x_show = item.get_attribute("x-show") or ""
            if "display: none" in style or "Pouces" in x_show:
                continue
            valeurs.append(item.inner_text().strip())
        if len(valeurs) >= 4:
            lignes.append({
                "taille_marque": valeurs[0],
                "taille_eu":     valeurs[0],
                "taille_uk":     valeurs[1],
                "taille_us":     valeurs[2],
                "taille_it":     None,
                "longueur_cm":   valeurs[3],
            })
        elif len(valeurs) == 3:
            lignes.append({
                "taille_marque": valeurs[0],
                "taille_eu":     valeurs[0],
                "taille_uk":     valeurs[1],
                "taille_us":     valeurs[2],
                "taille_it":     None,
                "longueur_cm":   None,
            })
    return lignes, "Kleman"

# ─────────────────────────────────────────────
# Scraper LA BOTTEGA GARDIANE
# ─────────────────────────────────────────────
def scraper_guide_gardiane(page, gender="Homme"):
    print("  Ouverture du guide de taille La Bottega Gardiane...")
    try:
        page.click("[data-trigger-size-guide]", timeout=10000)
        time.sleep(3)
        print("  Guide ouvert !")
    except Exception:
        print("  Bouton guide introuvable")
        return [], None

    try:
        page.wait_for_selector(".size-guide__table-right-col", timeout=8000)
    except Exception:
        print("  Tableau pas trouve")
        return [], None

    # Choisir le bon tableau selon le gender
    titre_cherche = "POINTURES HOMME" if gender == "Homme" else "POINTURES FEMME"
    print(f"  Lecture tableau : {titre_cherche}")

    # Trouver tous les tableaux dans le premier slide (slide01)
    tableaux = page.query_selector_all("#splide05-slide01 .size-guide__table")
    tableau_cible = None

    for tableau in tableaux:
        try:
            header = tableau.query_selector(".size-guide__table-cell.is--header")
            if header:
                header_text = header.inner_text().strip().upper()
                if titre_cherche in header_text:
                    tableau_cible = tableau
                    break
        except Exception:
            continue

    # Fallback : prendre tous les tableaux de la page
    if not tableau_cible:
        tous_tableaux = page.query_selector_all(".size-guide__table")
        for tableau in tous_tableaux:
            try:
                header = tableau.query_selector(".size-guide__table-cell.is--header")
                if header:
                    header_text = header.inner_text().strip().upper()
                    if titre_cherche in header_text:
                        tableau_cible = tableau
                        break
            except Exception:
                continue

    if not tableau_cible:
        print("  Tableau cible introuvable, on prend le premier")
        tous = page.query_selector_all(".size-guide__table")
        tableau_cible = tous[0] if tous else None

    if not tableau_cible:
        return [], None

    # Lire les colonnes (chaque colonne = une taille)
    # Structure : table-left contient [vide, FR, UK, US, IT]
    #             table-right-col contient [CM, FR_val, UK_val, US_val, IT_val]
    colonnes = tableau_cible.query_selector_all(".size-guide__table-right-col")
    lignes = []

    for col in colonnes:
        cells = col.query_selector_all(".size-guide__table-cell")
        valeurs = [c.inner_text().strip().replace("cm", "").strip() for c in cells]
        # valeurs = [CM, FR, UK, US, IT]
        if len(valeurs) >= 5:
            cm_val = valeurs[0].replace(",", ".") if valeurs[0] else None
            fr_val = valeurs[1].replace(",", ".") if valeurs[1] else None
            uk_val = valeurs[2].replace(",", ".") if valeurs[2] else None
            us_val = valeurs[3].replace(",", ".") if valeurs[3] else None
            it_val = valeurs[4].replace(",", ".") if valeurs[4] else None
            if fr_val and fr_val != "-":
                lignes.append({
                    "taille_marque": fr_val,
                    "taille_eu":     fr_val,
                    "taille_uk":     uk_val,
                    "taille_us":     us_val,
                    "taille_it":     it_val,
                    "longueur_cm":   cm_val,
                })

    return lignes, "La Bottega Gardiane"

# ─────────────────────────────────────────────
# Export Excel format horizontal
# ─────────────────────────────────────────────
def get_prochain_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == "Guide de taille" and row[1] and isinstance(row[1], int):
            max_id = max(max_id, row[1])
    return max_id + 1

def exporter_vers_excel(lignes, url, brand, guide_id):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Pages produit"
        wb.create_sheet("Guides de taille")
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    if "Guides de taille" not in wb.sheetnames:
        wb.create_sheet("Guides de taille")
    ws = wb["Guides de taille"]

    next_row = 1
    for row in ws.iter_rows():
        if any(cell.value is not None for cell in row):
            next_row = row[0].row + 1
    if next_row > 1:
        next_row += 1

    center     = Alignment(horizontal="center", vertical="center")
    blue_fill  = PatternFill("solid", start_color="00B0F0")
    dark_fill  = PatternFill("solid", start_color="1F4E79")
    cyan_fill  = PatternFill("solid", start_color="00FFFF")
    white_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    bold_font  = Font(bold=True, name="Arial", size=10)
    norm_font  = Font(name="Arial", size=10)

    # Ligne 1 : Guide de taille | ID | URL
    ws.cell(row=next_row, column=1, value="Guide de taille").font = white_font
    ws.cell(row=next_row, column=1).fill = dark_fill
    ws.cell(row=next_row, column=1).alignment = center
    ws.cell(row=next_row, column=2, value=guide_id).font = bold_font
    ws.cell(row=next_row, column=2).fill = cyan_fill
    ws.cell(row=next_row, column=2).alignment = center
    ws.cell(row=next_row, column=3, value="URL").font = bold_font
    ws.cell(row=next_row, column=3).alignment = center
    ws.cell(row=next_row, column=4, value=url).font = norm_font

    # Ligne 3 : headers
    row_h = next_row + 2
    ws.cell(row=row_h, column=1, value="Systemes metriques").font = white_font
    ws.cell(row=row_h, column=1).fill = dark_fill
    ws.cell(row=row_h, column=1).alignment = center
    ws.cell(row=row_h, column=2, value="").fill = blue_fill
    for i in range(len(lignes)):
        cell = ws.cell(row=row_h, column=i + 3, value=f"Taille {i+1}")
        cell.font = white_font
        cell.fill = dark_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(i + 3)].width = 10

    # Determiner les lignes de donnees selon la brand
    has_it = any(l.get("taille_it") for l in lignes)
    rows_data = [
        (brand,         brand, "taille_marque"),
        ("Europe",      "EU",  "taille_eu"),
        ("Royaume-Uni", "UK",  "taille_uk"),
        ("Etats-Unis",  "US",  "taille_us"),
    ]
    if has_it:
        rows_data.append(("Italie", "IT", "taille_it"))

    for offset, (label, short, key) in enumerate(rows_data, 1):
        r = row_h + offset
        ws.cell(row=r, column=1, value=label).font = norm_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2, value=short).font = white_font
        ws.cell(row=r, column=2).fill = blue_fill
        ws.cell(row=r, column=2).alignment = center
        for i, ligne in enumerate(lignes):
            cell = ws.cell(row=r, column=i + 3, value=ligne.get(key))
            cell.font = norm_font
            cell.alignment = center

    # Longueur pied
    r_cm = row_h + len(rows_data) + 1
    ws.cell(row=r_cm, column=1, value="Longueur pied").font = norm_font
    ws.cell(row=r_cm, column=1).alignment = center
    for i, ligne in enumerate(lignes):
        val = f"{ligne['longueur_cm']} cm" if ligne.get("longueur_cm") else None
        cell = ws.cell(row=r_cm, column=i + 3, value=val)
        cell.font = norm_font
        cell.alignment = center

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50

    wb.save(EXCEL_FILE)
    print(f"\n  {len(lignes)} tailles exportees dans '{EXCEL_FILE}' (ID guide = {guide_id})")

# ─────────────────────────────────────────────
# Programme principal
# ─────────────────────────────────────────────
def scrape_guide_taille(url, gender="Homme"):
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="fr-FR",
        )
        page = context.new_page()
        page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)

        print(f"\n Chargement de la page...")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        print(" Attente du chargement JavaScript...")
        time.sleep(4)
        accepter_cookies(page)
        time.sleep(2)

        if "prada.com" in url:
            lignes, brand = scraper_guide_prada(page)
        elif "kleman" in url:
            lignes, brand = scraper_guide_kleman(page, gender)
        elif "labottegardiane" in url or "bottega" in url:
            lignes, brand = scraper_guide_gardiane(page, gender)
        else:
            print("  Site non supporte (Prada, Kleman, La Bottega Gardiane supportes)")
            lignes, brand = [], None

        browser.close()
    return lignes, brand

def afficher_resultats(lignes, brand):
    if not lignes:
        print("\n Aucun guide de taille trouve\n")
        return
    has_it = any(l.get("taille_it") for l in lignes)
    print(f"\n {len(lignes)} tailles trouvees ({brand}) :")
    print("=" * 75)
    if has_it:
        print(f"  {'Marque':<10} {'EU/FR':<8} {'UK':<8} {'US':<8} {'IT':<8} {'cm':<8}")
    else:
        print(f"  {'Marque':<10} {'EU':<8} {'UK':<10} {'US':<10} {'cm':<10}")
    print("-" * 75)
    for l in lignes:
        if has_it:
            print(f"  {str(l['taille_marque']):<10} {str(l['taille_eu'] or '-'):<8} {str(l['taille_uk'] or '-'):<8} {str(l['taille_us'] or '-'):<8} {str(l['taille_it'] or '-'):<8} {str(l['longueur_cm'] or '-'):<8}")
        else:
            print(f"  {str(l['taille_marque']):<10} {str(l['taille_eu'] or '-'):<8} {str(l['taille_uk'] or '-'):<10} {str(l['taille_us'] or '-'):<10} {str(l['longueur_cm'] or '-'):<10}")
    print("=" * 75 + "\n")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("\nUsage : python scraper_guide_taille.py <URL> [Homme|Femme]")
        print("Exemples :")
        print("  python scraper_guide_taille.py https://www.prada.com/...")
        print("  python scraper_guide_taille.py https://kleman-france.com/... Homme")
        print("  python scraper_guide_taille.py https://www.labottegardiane.com/... Femme\n")
        sys.exit(1)

    url    = sys.argv[1]
    gender = sys.argv[2] if len(sys.argv) > 2 else "Homme"

    lignes, brand = scrape_guide_taille(url, gender)
    afficher_resultats(lignes, brand)

    if lignes:
        wb = load_workbook(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else None
        if wb and "Guides de taille" in wb.sheetnames:
            guide_id = get_prochain_id(wb["Guides de taille"])
        else:
            guide_id = 1
        exporter_vers_excel(lignes, url, brand, guide_id)
        print(f" Ouvre '{EXCEL_FILE}' pour voir le resultat !\n")
