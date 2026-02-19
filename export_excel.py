"""
Export Excel - Onglet 1 : Pages produit
Usage : python export_excel.py <URL>
Exemple : python export_excel.py https://kleman-france.com/products/padror-th-cognac
"""

import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Importer le scraper qu'on a deja fait
from scraper_produit import scrape_product

EXCEL_FILE = "etudes_de_cas.xlsx"


def creer_fichier_excel():
    """Cree un nouveau fichier Excel avec les 2 onglets."""
    wb = Workbook()

    # ── Onglet 1 : Pages produit ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Pages produit"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    headers = ["Nom Produit", "Gender", "Type", "URL", "Guide de taille"]
    col_widths = [40, 12, 12, 55, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.row_dimensions[1].height = 20

    # ── Onglet 2 : Guides de taille ───────────────────────────
    ws2 = wb.create_sheet("Guides de taille")

    headers2 = ["ID", "Nom Produit", "Brand", "Type", "URL Source", "Taille FR", "Taille IT", "Taille UK", "Taille US", "Longueur (cm)"]
    col_widths2 = [8, 35, 15, 12, 50, 12, 12, 12, 12, 14]

    for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.row_dimensions[1].height = 20

    wb.save(EXCEL_FILE)
    print(f" Fichier '{EXCEL_FILE}' cree avec succes !")
    return wb


def ajouter_produit(data: dict):
    """Ajoute une ligne dans l'onglet 1 du Excel."""

    # Creer le fichier s'il n'existe pas encore
    if not os.path.exists(EXCEL_FILE):
        creer_fichier_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Pages produit"]

    # Trouver la prochaine ligne vide
    next_row = ws.max_row + 1

    # Alterner les couleurs de lignes
    if next_row % 2 == 0:
        row_fill = PatternFill("solid", start_color="D6E4F0")
    else:
        row_fill = PatternFill("solid", start_color="FFFFFF")

    normal_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")

    # Ecrire les donnees
    valeurs = [
        data.get("titre"),
        data.get("gender"),
        data.get("type"),
        data.get("url"),
        None,  # Guide de taille - sera rempli plus tard
    ]

    for col, val in enumerate(valeurs, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = normal_font
        cell.fill = row_fill
        cell.alignment = center if col in [2, 3, 5] else Alignment(vertical="center")

    wb.save(EXCEL_FILE)
    print(f" Produit ajoute a la ligne {next_row} dans '{EXCEL_FILE}' !")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("\nErreur : tu dois fournir une URL en argument.")
        print("Usage : python export_excel.py <URL>\n")
        sys.exit(1)

    url = sys.argv[1]

    if not url.startswith("http"):
        print("\nErreur : l'URL doit commencer par http:// ou https://\n")
        sys.exit(1)

    print(f"\n Scraping du produit...")
    data = scrape_product(url)

    print(f"\n Resultats trouves :")
    print(f"  Titre  : {data['titre']  or 'Non trouve'}")
    print(f"  Gender : {data['gender'] or 'Non trouve'}")
    print(f"  Type   : {data['type']   or 'Non trouve'}")
    print(f"  URL    : {data['url']}")

    ajouter_produit(data)
    print(f"\n Ouvre '{EXCEL_FILE}' pour voir le resultat !\n")
